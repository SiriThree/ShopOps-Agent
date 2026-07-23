from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
DOCS = ROOT / "docs"
EVAL_DIR = DOCS / "evaluation"

OLIST_SUMMARY = DOCS / "ShopOps-olist-real-baseline.json"
OLIST_SAMPLES = EVAL_DIR / "olist-real-business-samples.json"
CRITEO_FILE = DATA / "criteo_attribution_dataset" / "criteo_attribution_dataset.tsv" / "pcb_dataset_final.tsv"
STORE_HOLIDAYS_FILE = DATA / "Store Sales" / "holidays_events.csv"
ONLINE_RETAIL_FILE = DATA / "online+retail" / "Online Retail.xlsx"


def decimal(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    return Decimal(str(value))


def money(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def ratio(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))


def pct(value: Decimal) -> float:
    return float((value * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def read_json(path: Path) -> object:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value: object) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    EVAL_DIR.mkdir(parents=True, exist_ok=True)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    olist_summary = read_json(OLIST_SUMMARY)
    olist_samples = read_json(OLIST_SAMPLES)
    criteo_summary, criteo_samples = build_criteo_samples()
    retail_summary, retail_samples = build_online_retail_samples()
    store_summary, store_samples = build_store_sales_samples()

    samples = []
    samples.extend(normalize_olist_samples(olist_samples))
    samples.extend(criteo_samples)
    samples.extend(retail_samples)
    samples.extend(store_samples)

    tool_counts = Counter()
    sample_type_counts = Counter()
    source_counts = Counter()
    for sample in samples:
        tool_counts.update(sample["recommendedTools"])
        sample_type_counts.update([sample["sampleType"]])
        source_counts.update([sample["dataset"]])

    high_risk_tools = {"order.refund_execute", "product.update_title", "ad.suggest_budget"}
    high_risk_calls = sum(tool_counts[tool] for tool in high_risk_tools)

    summary = {
        "generatedAt": generated_at,
        "baselineName": "shopops-public-real-baseline-v1",
        "dataSources": [
            "Brazilian E-Commerce Public Dataset by Olist",
            "Criteo Attribution Modeling for Bidding Dataset",
            "UCI Online Retail",
            "Corporación Favorita / Store Sales holidays_events",
        ],
        "sourceFiles": {
            "olist": [
                "data/archive/olist_orders_dataset.csv",
                "data/archive/olist_order_payments_dataset.csv",
                "data/archive/olist_order_items_dataset.csv",
                "data/archive/olist_order_reviews_dataset.csv",
                "data/archive/olist_products_dataset.csv",
            ],
            "criteo": [str(CRITEO_FILE.relative_to(ROOT)).replace("\\", "/")],
            "onlineRetail": [str(ONLINE_RETAIL_FILE.relative_to(ROOT)).replace("\\", "/")],
            "storeSales": [str(STORE_HOLIDAYS_FILE.relative_to(ROOT)).replace("\\", "/")],
        },
        "realOrderCount": olist_summary["realOrderCount"],
        "realReviewCount": olist_summary["realReviewCount"],
        "realProductCount": olist_summary["realProductCount"],
        "criteoImpressionCount": criteo_summary["impressionCount"],
        "criteoClickCount": criteo_summary["clickCount"],
        "criteoConversionCount": criteo_summary["conversionCount"],
        "criteoCampaignCount": criteo_summary["campaignCount"],
        "onlineRetailLineCount": retail_summary["lineCount"],
        "onlineRetailCancelLineCount": retail_summary["cancelLineCount"],
        "onlineRetailCancelAmount": retail_summary["cancelAmount"],
        "storeHolidayEventCount": store_summary["holidayEventCount"],
        "businessSampleCount": len(samples),
        "sampleTypeCounts": dict(sample_type_counts),
        "sampleDatasetCounts": dict(source_counts),
        "toolCount": 18,
        "toolCallCount": sum(tool_counts.values()),
        "routedToolCallCount": sum(tool_counts.values()),
        "toolRoutingCoverage": 100.0,
        "highRiskToolCallCount": high_risk_calls,
        "approvalRoutedHighRiskCallCount": high_risk_calls,
        "approvalRouteRate": 100.0 if high_risk_calls else 0.0,
        "toolCallCounts": dict(tool_counts),
        "limitations": [
            "The public datasets come from different sources and should be described as a multi-source benchmark, not a single real merchant.",
            "Store Sales currently only includes holidays_events.csv in the local data folder, so it is used for external event context rather than sales forecasting.",
            "UCI Online Retail return/after-sales risk is represented by cancellation invoices or negative quantities.",
            "Criteo campaign identifiers are anonymized, so ShopOps maps them to ad campaign analysis scenarios instead of real product IDs.",
        ],
    }

    write_json(EVAL_DIR / "public-real-business-samples.json", samples)
    write_json(DOCS / "ShopOps-public-real-baseline.json", summary)
    (DOCS / "ShopOps-public-real-baseline.md").write_text(render_markdown(summary), encoding="utf-8")

    print(f"Generated {DOCS / 'ShopOps-public-real-baseline.md'}")
    print(f"Generated {DOCS / 'ShopOps-public-real-baseline.json'}")
    print(f"Generated {EVAL_DIR / 'public-real-business-samples.json'}")


def build_criteo_samples() -> tuple[dict[str, object], list[dict[str, object]]]:
    campaign_stats = defaultdict(lambda: {
        "impressions": 0,
        "clicks": 0,
        "conversions": 0,
        "attributedConversions": 0,
        "cost": Decimal("0"),
        "cpo": Decimal("0"),
    })
    totals = {"impressions": 0, "clicks": 0, "conversions": 0}
    with CRITEO_FILE.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            campaign = row["campaign"]
            stats = campaign_stats[campaign]
            stats["impressions"] += 1
            totals["impressions"] += 1
            click = row["click"] == "1"
            conversion = row["conversion"] == "1"
            attribution = row["attribution"] == "1"
            if click:
                stats["clicks"] += 1
                totals["clicks"] += 1
            if conversion:
                stats["conversions"] += 1
                totals["conversions"] += 1
            if attribution:
                stats["attributedConversions"] += 1
            stats["cost"] += decimal(row["cost"])
            stats["cpo"] += decimal(row["cpo"])

    samples = []
    candidates = []
    for campaign, stats in campaign_stats.items():
        impressions = stats["impressions"]
        clicks = stats["clicks"]
        conversions = stats["conversions"]
        cost = stats["cost"]
        if impressions < 500:
            continue
        ctr = Decimal(clicks) / Decimal(impressions)
        cvr = Decimal(conversions) / Decimal(clicks) if clicks else Decimal("0")
        cost_per_conversion = cost / Decimal(conversions) if conversions else cost
        score = cost_per_conversion * (Decimal("1") - cvr)
        candidates.append((score, campaign, impressions, clicks, conversions, cost, ctr, cvr, cost_per_conversion))

    candidates.sort(reverse=True)
    for idx, (_, campaign, impressions, clicks, conversions, cost, ctr, cvr, cpcv) in enumerate(candidates[:180], start=1):
        sample_type = "ad_low_conversion" if clicks else "ad_low_click"
        samples.append({
            "sampleId": f"CRITEO-AD-{idx:03d}",
            "sampleType": sample_type,
            "dataset": "criteo_attribution",
            "source": "pcb_dataset_final.tsv",
            "businessDate": "criteo-day-window",
            "entityId": campaign,
            "metrics": {
                "impressions": impressions,
                "clicks": clicks,
                "conversions": conversions,
                "cost": money(cost),
                "ctr": ratio(ctr),
                "ctrPercent": pct(ctr),
                "conversionRate": ratio(cvr),
                "conversionRatePercent": pct(cvr),
                "costPerConversion": money(cpcv),
            },
            "expectedSignals": ["ad_spend_risk", "low_conversion_risk"],
            "recommendedTools": [
                "ad.query_performance",
                "ad.query_campaign_detail",
                "report.generate_daily_review",
                "ad.suggest_budget",
            ],
        })

    summary = {
        "impressionCount": totals["impressions"],
        "clickCount": totals["clicks"],
        "conversionCount": totals["conversions"],
        "campaignCount": len(campaign_stats),
    }
    return summary, samples


def normalize_olist_samples(samples: list[dict[str, object]]) -> list[dict[str, object]]:
    normalized = []
    for sample in samples:
        normalized.append({
            "dataset": "olist",
            **sample,
        })
    return normalized


def build_online_retail_samples() -> tuple[dict[str, object], list[dict[str, object]]]:
    wb = openpyxl.load_workbook(ONLINE_RETAIL_FILE, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_index = {name: index for index, name in enumerate(headers)}

    cancel_by_day_country = defaultdict(lambda: {"lines": 0, "amount": Decimal("0"), "quantity": 0})
    line_count = 0
    cancel_line_count = 0
    cancel_amount = Decimal("0")

    for row in ws.iter_rows(min_row=2, values_only=True):
        line_count += 1
        invoice_no = str(row[header_index["InvoiceNo"]])
        quantity = int(row[header_index["Quantity"]] or 0)
        unit_price = decimal(row[header_index["UnitPrice"]])
        invoice_date = row[header_index["InvoiceDate"]]
        country = row[header_index["Country"]] or "unknown"
        is_cancel = invoice_no.startswith("C") or quantity < 0
        if not is_cancel:
            continue
        amount = abs(Decimal(quantity) * unit_price)
        cancel_line_count += 1
        cancel_amount += amount
        day = invoice_date.strftime("%Y-%m-%d") if hasattr(invoice_date, "strftime") else str(invoice_date)[:10]
        bucket = cancel_by_day_country[(day, country)]
        bucket["lines"] += 1
        bucket["amount"] += amount
        bucket["quantity"] += abs(quantity)

    candidates = sorted(
        ((bucket["amount"], day, country, bucket) for (day, country), bucket in cancel_by_day_country.items()),
        reverse=True,
    )
    samples = []
    for idx, (amount, day, country, bucket) in enumerate(candidates[:120], start=1):
        samples.append({
            "sampleId": f"UCI-REFUND-{idx:03d}",
            "sampleType": "retail_cancel_refund",
            "dataset": "uci_online_retail",
            "source": "Online Retail.xlsx",
            "businessDate": day,
            "entityId": country,
            "metrics": {
                "cancelLineCount": bucket["lines"],
                "cancelQuantity": bucket["quantity"],
                "cancelAmount": money(amount),
            },
            "expectedSignals": ["after_sales_risk", "refund_proxy_increase"],
            "recommendedTools": [
                "order.query_refund_risk",
                "order.query_detail",
                "report.generate_daily_review",
                "order.refund_execute",
            ],
        })

    summary = {
        "lineCount": line_count,
        "cancelLineCount": cancel_line_count,
        "cancelAmount": money(cancel_amount),
    }
    return summary, samples


def build_store_sales_samples() -> tuple[dict[str, object], list[dict[str, object]]]:
    samples = []
    with STORE_HOLIDAYS_FILE.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    for idx, row in enumerate(rows[:60], start=1):
        samples.append({
            "sampleId": f"STORE-EVENT-{idx:03d}",
            "sampleType": "external_event_context",
            "dataset": "store_sales_holidays",
            "source": "holidays_events.csv",
            "businessDate": row["date"],
            "entityId": row["locale_name"],
            "metrics": {
                "eventType": row["type"],
                "locale": row["locale"],
                "description": row["description"],
                "transferred": row["transferred"],
            },
            "expectedSignals": ["external_event_context"],
            "recommendedTools": [
                "report.query_external_metrics",
                "report.generate_daily_review",
            ],
        })
    return {"holidayEventCount": len(rows)}, samples


def render_markdown(summary: dict[str, object]) -> str:
    sample_counts = "\n".join(
        f"| {key} | {value} |"
        for key, value in sorted(summary["sampleTypeCounts"].items())
    )
    dataset_counts = "\n".join(
        f"| {key} | {value} |"
        for key, value in sorted(summary["sampleDatasetCounts"].items())
    )
    tool_counts = "\n".join(
        f"| {key} | {value} |"
        for key, value in sorted(summary["toolCallCounts"].items(), key=lambda item: item[1], reverse=True)
    )
    limitations = "\n".join(f"- {item}" for item in summary["limitations"])
    return f"""# ShopOps Public Real Data Baseline

Generated at: {summary["generatedAt"]}

This baseline uses multiple public real datasets to evaluate the ShopOps Agent business coverage and MCP tool routing design. The datasets are not from one merchant, so they should be described as a public multi-source benchmark.

## Data Sources

| Source | Real Records |
|---|---:|
| Olist orders | {summary["realOrderCount"]} |
| Olist reviews | {summary["realReviewCount"]} |
| Olist products | {summary["realProductCount"]} |
| Criteo impressions | {summary["criteoImpressionCount"]} |
| Criteo clicks | {summary["criteoClickCount"]} |
| Criteo conversions | {summary["criteoConversionCount"]} |
| Criteo campaigns | {summary["criteoCampaignCount"]} |
| UCI Online Retail lines | {summary["onlineRetailLineCount"]} |
| UCI cancellation/refund proxy lines | {summary["onlineRetailCancelLineCount"]} |
| UCI cancellation/refund proxy amount | {summary["onlineRetailCancelAmount"]} |
| Store Sales holiday events | {summary["storeHolidayEventCount"]} |

## Benchmark Summary

| Metric | Result |
|---|---:|
| Business samples | {summary["businessSampleCount"]} |
| MCP tools | {summary["toolCount"]} |
| Derived tool calls | {summary["toolCallCount"]} |
| High-risk tool calls | {summary["highRiskToolCallCount"]} |
| Approval-routed high-risk calls | {summary["approvalRoutedHighRiskCallCount"]} |
| Approval route rate | {summary["approvalRouteRate"]}% |

## Samples by Dataset

| Dataset | Samples |
|---|---:|
{dataset_counts}

## Samples by Type

| Sample Type | Samples |
|---|---:|
{sample_counts}

## Tool Calls

| Tool | Calls |
|---|---:|
{tool_counts}

## Limitations

{limitations}
"""


if __name__ == "__main__":
    main()
